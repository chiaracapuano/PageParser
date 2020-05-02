def rearrange_to_df(df, dim):
    return pd.DataFrame(np.array(df).reshape(int(len(df) / dim), dim))

def col_tot(df, col):
    return int(df[df[0]=='Total'][col])

def filter_illegals(table):
    illegals = ['†', '\n', '‡', 'StartFragment', 'EndFragment', '*/\n*/']
    return list(filter(lambda x: x not in illegals,table))


def col_converter(df):
    for col in df.columns:
        try:
            df[col] = df[col].astype('float')
        except ValueError:
            print("Moving to next column")


def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False,
                       **to_excel_kwargs):
    """
    A\[nd a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]

    Returns: None
    """
    from openpyxl import load_workbook

    import pandas as pd

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    # Python 2.x: define [FileNotFoundError] exception if it doesn't exist
    try:
        FileNotFoundError
    except NameError:
        FileNotFoundError = IOError


    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()



import bs4 as bs
import urllib.request
source = urllib.request.urlopen('https://govstatus.egov.com/OR-OHA-COVID-19')
soup = bs.BeautifulSoup(source, "html.parser")

#OR counties
tab= soup.find_all(text=True)
tab_i = tab.index("County")
tab_f = tab.index("Age group")
tab_ii = tab.index("Age group")
tab_ff = tab.index("Sex")
tab_age = tab[tab_ii:tab_ff]
demographics_table = tab[tab_i:tab_f]
demographics_table_updated = filter_illegals(demographics_table)

import pandas as pd
import numpy as np

dem_df = rearrange_to_df(demographics_table_updated, 4)
dem_df = dem_df[1:] #take the data less the header row

from datetime import date,timedelta
yesterday =  (date.today() - timedelta(days=1)).strftime('%d-%b-%y')
today = date.today().strftime('%d-%b-%y')
dem_df[3] = today

print(dem_df)
col_converter(dem_df)

#checking data quality
sum_col1  = col_tot(dem_df, 1)
sum_col2  = col_tot(dem_df, 2)

if int(dem_df[1].sum()/2) == sum_col1 and int(dem_df[2].sum()/2) == sum_col2:
    print("Appending file on OR counties")
 #   append_df_to_excel(r'C:\Users\adach\Desktop\covidOR.xlsx', dem_df, sheet_name='Sheet1', index=False,
  #                 header=None)

else:
    print("Check Counties manually")


tab_age_updated = filter_illegals(tab_age)
df = rearrange_to_df(tab_age_updated, 5)

df = df[1:]
df[3] = today
df = df.drop([4], axis=1)
col_converter(df)

list_add = pd.Series(['17 or younger','18 to 24','25 to 34','35 to 54','55+', 'Total'])
df_to_concat = list_add.to_frame()
df_to_concat[1]=""
df_to_concat[2]=""
df_to_concat[3] = today
ages_df = pd.concat([df_to_concat,df])
#append_df_to_excel(r'C:\Users\adach\Desktop\covidORages.xlsx', ages_df, sheet_name='Sheet1', index=False,
#                    header=None)
print(ages_df)
print("Appending file on OR demographics")

#Canada
source = urllib.request.urlopen('https://www.canada.ca/en/public-health/services/diseases/2019-novel-coronavirus-infection.html')
soup = bs.BeautifulSoup(source, "html.parser")

tab= soup.find_all(text=True)
tab_i = tab.index("Province, territory or other")
tab_f = tab.index('Additional COVID-19 case information:')
tab_ca = tab[tab_i:tab_f]
ca_table_updated = filter_illegals(tab_ca)

ca_df = rearrange_to_df(ca_table_updated, 4)
ca_df = ca_df[1:] #take the data less the header row
today = date.today().strftime('%d/%m/%y')

ca_df[2] = today
import locale
from locale import atof
locale.setlocale(locale.LC_NUMERIC, '')
ca_df[1] = ca_df[1].map(atof)
ca_df[3] = ca_df[3].map(atof)

col_converter(ca_df)
print(ca_df)
#append_df_to_excel(r'C:\Users\adach\Desktop\Canada-Covid.xlsx', ca_df, sheet_name='Sheet1', index=False,
 #                    header=None)

#USA
source = urllib.request.urlopen('https://raw.githubusercontent.com/COVID19Tracking/covid-tracking-data/master/data/states_daily_4pm_et.csv')
soup = bs.BeautifulSoup(source, "html.parser")
tab= soup.get_text()
import re
usa = re.split('\n|,', tab)
df_usa = rearrange_to_df(usa,25)
col_converter(df_usa)
df_usa.to_excel(r'C:\Users\adach\Desktop\states_daily_4pm_et.xlsx', header = None, index = False)

#ITA reg

source = urllib.request.urlopen('https://raw.githubusercontent.com/pcm-dpc/COVID-19/master/dati-regioni/dpc-covid19-ita-regioni.csv')
soup = bs.BeautifulSoup(source, "html.parser")
tab= soup.get_text()
import re
ita_reg = re.split('\n|,', tab)
df_ita_reg = rearrange_to_df(ita_reg[:-1],20)
col_converter(df_ita_reg)
df_ita_reg.to_excel(r'C:\Users\adach\Desktop\dpc-covid19-ita-regioni.xlsx', header = None, index = False)


#ITA prov

source = urllib.request.urlopen('https://raw.githubusercontent.com/pcm-dpc/COVID-19/master/dati-province/dpc-covid19-ita-province.csv')
soup = bs.BeautifulSoup(source, "html.parser")
tab= soup.get_text()
import re
ita_prov = re.split('\n|,', tab)
df_ita_prov = rearrange_to_df(ita_prov[:-1],12)
col_converter(df_ita_prov)
df_ita_prov.to_excel(r'C:\Users\adach\Desktop\dpc-covid19-ita-province.xlsx', header = None, index = False)


#ITA naz

source = urllib.request.urlopen('https://raw.githubusercontent.com/pcm-dpc/COVID-19/master/dati-andamento-nazionale/dpc-covid19-ita-andamento-nazionale.csv')
soup = bs.BeautifulSoup(source, "html.parser")
tab= soup.get_text()
import re
ita_nat = re.split('\n|,', tab)
print(ita_nat)
df_ita_nat = rearrange_to_df(ita_nat[:-2],16)

col_converter(df_ita_nat)

df_ita_nat.to_excel(r'C:\Users\adach\Desktop\dpc-covid19-ita-andamento-nazionale.xlsx', header = None, index = False)